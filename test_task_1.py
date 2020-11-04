#!/usr/bin/env python
# encoding:     UTF-8
# filename:     test_task_1.py
# project:      TEST_TASK_1
script_ver = "1.0.3"
script_date = "2020.11.04"
print(f"Version=[{script_ver}] Date=[{script_date}]")
# company:
# author:       Starichenko Andrey
# language:     Python 3.8.6rc1 (AMD64)
import sys; print(sys.version)
# OS:           Windows 8.1 SL RU (AMD64)
import platform; print(platform.platform(), platform.machine())
# IDE:          PyCharm Community 2020.2
# licence: 	    GPLv3

# todo:1: check ID for unique! in all column at start!!!

# TODO:0: implement AUTOTEST!
# TODO:0: use PATHLIB for names


import os
import time
import glob
import re
from openpyxl import load_workbook, Workbook    # pip install openpyxl
import timeit
import winsound
from threading import Thread
import threading

SOURCE_FILE_NAME = "ДАННЫЕ.xlsx"

# PATH FORMAT: without starting and finishing slashes!
PATH1 = "Папка1"
PATH2 = "Папка2"
PATH3 = "Папка3"

# DATE FORMAT: "dd.mm.yyyy"!
DATE_START = "20.06.2020"
DATE_FINISH = "10.07.2020"


date_start_t = time.strptime(DATE_START, "%d.%m.%Y")
date_finish_t = time.strptime(DATE_FINISH, "%d.%m.%Y")


ERROR_LIST = [
    ["ID", "MSG_PRINT"],
    [1, "[ERROR/filenames]CAN\'T SAVE FILE, check SYMBOLS IN TEXT!!!"],
    [2, "[ERROR/dateOrOther] may be incorrect date (or other unknown error)"],
]


def main():
    if not os.path.exists(SOURCE_FILE_NAME):
        print(f"ERROR: source file not exists! [{SOURCE_FILE_NAME}]")
        return

    make_directories()

    wb_defects = Workbook()
    ws_defects = wb_defects.active

    # todo:1: check ID for unique! in all column at start!!!
    for source_line in read_source_lines(SOURCE_FILE_NAME):
        defected_line = False

        try:
            for filename in make_filename_group(source_line):
                for pathname in filter_and_make_path_group(source_line):
                    try:
                        make_file(pathname, filename)
                    except:
                        error_i = ERROR_LIST[1]
                        print(error_i[1], pathname, filename)
                        defected_line = True

                if defected_line:
                    break #stop listing "for filename"

        except:
            error_i = ERROR_LIST[2]
            print(error_i[1], source_line)
            defected_line = True

        if defected_line:
            delete_files_for_defected_row(source_line)
            ws_defects.append(list(source_line)+[error_i[1]])

            alert_error()

    if list(ws_defects.values) != []:
        defected_xls_filename_nameonly = "DEFECTED_DATA=check manually.xlsx"
        defected_xls_filename = PATH_RESULT + defected_xls_filename_nameonly

        max_defected_rows = ws_defects.max_row
        print("*"*80)
        print(f"количество дефектных строк=[{max_defected_rows}]")
        wb_defects.save(defected_xls_filename)
        wb_defects.close()

        make_copy_this_script_to_result_folder(defected_xls_filename_nameonly)


def make_directories():
    """make directories for saving results"""
    global PATH1, PATH2, PATH3, PATH_RESULT
    PATH_RESULT = "РЕЗУЛЬТАТЫ_выгрузки_" + time.strftime("%Y.%m.%d_%H.%M.%S", time.localtime()) + "/"
    PATH1 = PATH_RESULT + PATH1
    PATH2 = PATH_RESULT + PATH2
    PATH3 = PATH_RESULT + PATH3

    os.mkdir(PATH_RESULT)
    os.mkdir(PATH1)
    os.mkdir(PATH2)
    os.mkdir(PATH3)


def _filter1(row_tuple):  # D-column
    return str(row_tuple[3]) == "1"


def _filter2(row_tuple):  # last 2 digits in CA
    return row_tuple[0][-2] == row_tuple[0][-1]


def _filter3(row_tuple):  # date
    date_doc_t = time.strptime(row_tuple[2], "%d.%m.%Y")
    return date_start_t < date_doc_t < date_finish_t


def read_source_lines(filename):
    """
    Process: from Excel file read lines.
    Input: Excel filename(str)  ="Задание.xlsx".
    Output: rows (iter)         =<generator object Worksheet._cells_by_row at 0x000000D23AF32270>
    """
    wb = load_workbook(filename)
    ws = wb.active
    iter = ws.iter_rows(values_only=True)
    return iter


def filter_and_make_path_group(source_line):
    """
    1. Filter actual source line data from Excel
    2. Make corresponding list of path where will actually NEED saving files.
    Input: actual Excel row (tuple)         =('CA_0001', 'акт,счет', '30.06.2020', 1)
    Output: path list to save files (list)  =['РЕЗУЛЬТАТЫ_выгрузки_2020.10.06_17.58.48/Папка1', 'РЕЗУЛЬТАТЫ_выгрузки_2020.10.06_17.58.48/Папка3']
    """
    output = []

    if _filter1(source_line): output.append(PATH1)
    if _filter2(source_line): output.append(PATH2)
    if _filter3(source_line): output.append(PATH3)
    return output


def make_filename_group(row_tuple):
    """
    in relation to data in 2nd column make filenames need to save
    Input: actual Excel row (tuple)     =('CA_0001', 'акт,счет', '30.06.2020', 1)
    Output: filename list (list)        =['КА_CA_0001_акт_30.06.2020.txt', 'КА_CA_0001_счет_30.06.2020.txt']
    """
    output = []

    if row_tuple[0] in ['Код_КА', ""]:
        pass
    else:
        source_list_copy = list(row_tuple)
        for i in row_tuple[1].split(","):
            source_list_copy[1] = i
            output.append(_make_filename(source_list_copy))
    return output


def _make_filename(cells_data):
    """
    make filename only from Excel cells by template
    Input: isolated data for one file (list)    =['CA_0001', 'акт', '30.06.2020', 1]
    Output: filename (str)                      ="КА_CA_0001_акт_30.06.2020.txt"
    """
    if cells_data == []: return None

    part1_ca = cells_data[0]
    part2_type = cells_data[1]
    part3_date = cells_data[2]
    # make filename ("КА_<Код КА>_<тип_документа>_<дата>.txt")
    file_name_i = f"КА_{part1_ca}_{part2_type}_{part3_date}.txt"
    return file_name_i


def make_file(pathname, filename):
    """
    actually make file
    Input: path (str) and filename (str)    ="РЕЗУЛЬТАТЫ_выгрузки_2020.10.06_18.16.48/Папка1", "КА_CA_0001_акт_30.06.2020.txt"
    Output: save file (None)                =
    """
    if pathname in [None, ""] or filename in [None, ""]:
        pass
    else:
        full_filename = f"{pathname}/{filename}"
        newfile = open(full_filename, "w")
        newfile.close()
    return


def delete_files_for_defected_row(source_line):
    """
    if get error trying save any file, there are possibly already saved previous file in group.
    delete all group!
    Input: actual Excel row (tuple)     =('CA_0001', 'акт,счет', '30.06.2020', 1)
    Output: delete files (None)         =
    """
    delete_files_list = glob.glob(f"{PATH_RESULT}*/*{source_line[0]}_*.txt")
    for i in delete_files_list:
        os.remove(i)


def make_copy_this_script_to_result_folder(defected_xls_filename_nameonly):
    """
    make copy this script to result folder near the defected data excel file
    and chance link there to process it in next time.
    so you can manually edit defected excel file and start process again in the same folder.
    Input: filename (str)               ="DEFECTED_DATA=check manually.xlsx"
    Output: make edited copy (None)     =
    """
    this_script_copy_dstname = PATH_RESULT + os.path.basename(__file__)

    fp_this_script = open(__file__, "r", encoding="utf-8")
    fp_this_script_copy = open(this_script_copy_dstname, "w", encoding="utf-8")

    for line in fp_this_script:
        pattern = "SOURCE_FILE_NAME\s*=\s*.*[.]xls.*"
        repl = f'SOURCE_FILE_NAME = "{defected_xls_filename_nameonly}"\n'
        new_line = re.sub(pattern, repl, line)
        fp_this_script_copy.write(new_line)

    fp_this_script.close()
    fp_this_script_copy.close()


def alert_error():
    if threading.active_count() > 1:
        return

    try:
        def play_my_sound():
            winsound.Beep(1000, 500)

        th = Thread(target=play_my_sound)
        th.start()
    except:
        pass


if __name__ == '__main__':
    print("mainSTART" + "*" * 70)
    time_of_process = timeit.timeit(stmt="main()", number=1, globals=globals())
    print("mainFINISH" + "*"*70)
    print(f"Время выполнения: [{time_of_process}]секунд")
    print("!!!основной ПРОЦЕСС ЗАВЕРШЕН!!!")

